import os
import openpyxl

def main():
    path = r"C:\Users\jubam\Downloads\RQ4_Evaluation_Responses.xlsx"
    if not os.path.exists(path):
        print(f"Error: File not found at {path}")
        return
        
    print(f"Opening Excel file: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)
    for sheet_name in wb.sheetnames:
        print(f"\n--- Sheet: {sheet_name} ---")
        sheet = wb[sheet_name]
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if row_idx < 15:  # Print first 15 rows
                print(row)
                
if __name__ == '__main__':
    main()
